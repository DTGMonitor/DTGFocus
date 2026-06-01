"""
data_loader.py — File validation, sheet parsing, and VCP extraction.

Public interface
----------------
:func:`load_excel` is the sole public entry point.  It accepts raw file bytes
(as returned by ``st.file_uploader``), validates the workbook, parses all four
required sheets, and returns a :class:`~models.LoadResult`.

MonitorQ Excel structure (derived from Enhanced Deformation.xlsx)
-----------------------------------------------------------------
Sheet 0  EnhancedDeformation  — Time (Excel serial float) + displacement +
                                 velocity / inverse-velocity / acceleration
                                 columns for multiple window sizes.
Sheet 1  V-IV Charts          — Chart placeholder (may be empty).
Sheet 2  Acc Charts           — Chart placeholder (may be empty).
Sheet 3  Summary              — Progressive start times, notification times,
                                 failure times per window size.

VCP names are derived from the raw-data sheet name (Sheet 0).  In a
single-VCP workbook the list contains exactly one entry.
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta

import openpyxl
import pandas as pd

from models import LoadResult

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Expected number of sheets in a valid MonitorQ export.
_REQUIRED_SHEET_COUNT = 4

# Human-readable names used in error messages (index → label).
_SHEET_LABELS = {
    0: "EnhancedDeformation (Sheet 1)",
    1: "V-IV Charts (Sheet 2)",
    2: "Acc Charts (Sheet 3)",
    3: "Summary (Sheet 4)",
}

# Required columns in Sheet 1 (case-insensitive prefix match for displacement).
# Accepted column names for the time column (case-insensitive exact match,
# whitespace and punctuation collapsed).  The first hit wins.
_TIME_COLUMN_ALIASES: tuple[str, ...] = (
    "time",
    "timestamp",
    "datetime",
    "date time",
    "date/time",
    "date_time",
    "date",
)


def _normalise_header(name: object) -> str:
    """Collapse whitespace / punctuation in a header for matching.

    e.g. 'Date / Time' -> 'date time', 'date_time' -> 'date time'.
    """
    if not isinstance(name, str):
        return ""
    s = name.strip().lower()
    # Treat /, _, and any run of whitespace as a single space.
    import re as _re
    s = _re.sub(r"[\s/_]+", " ", s)
    return s.strip()
_DISPLACEMENT_COLUMN_KEYWORD = "Enhanced Deformation"

# Minimum fraction of non-null values needed to treat a column as a VCP.
_VCP_MIN_FILL_RATIO: float = 0.10

# Excel epoch: serial date 1 corresponds to 1900-01-01 in Excel's convention.
# Excel incorrectly treats 1900 as a leap year, so serial 1 = 1900-01-01 and
# serial 60 is the phantom 1900-02-29.  The standard correction is to use
# 1899-12-30 as the epoch (which absorbs the off-by-two).
_EXCEL_EPOCH = datetime(1899, 12, 30)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _excel_serial_to_datetime(serial: float) -> pd.Timestamp:
    """Convert an Excel serial date (float) to a :class:`pd.Timestamp`.

    Excel stores dates as the number of days since 1899-12-30 (with the
    fractional part representing the time of day).
    """
    return pd.Timestamp(_EXCEL_EPOCH + timedelta(days=serial))


def _is_sheet_empty(ws) -> bool:
    """Return True if a worksheet contains no non-None values."""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                return False
    return True


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------


def load_excel(
    file_bytes: bytes,
    max_size_bytes: int = 50 * 1024 * 1024,
) -> LoadResult:
    """Validate, parse, and extract data from a MonitorQ Excel export.

    Parameters
    ----------
    file_bytes:
        Raw bytes of the uploaded ``.xlsx`` / ``.xls`` file.
    max_size_bytes:
        Maximum accepted file size in bytes.  Defaults to 50 MB.

    Returns
    -------
    LoadResult
        Populated result object.  Check ``result.errors`` for fatal issues
        before using ``result.raw_df`` or ``result.vcp_names``.
    """
    result = LoadResult()

    # ------------------------------------------------------------------
    # Step 1: Validate file size.
    # ------------------------------------------------------------------
    actual_size = len(file_bytes)
    if actual_size > max_size_bytes:
        result.errors.append(
            f"Data Loader: File size {actual_size:,} bytes exceeds the "
            f"{max_size_bytes:,}-byte (50 MB) limit.  Please upload a "
            f"smaller file."
        )
        return result

    # ------------------------------------------------------------------
    # Step 2: Attempt to open as an Excel workbook.
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        result.errors.append(
            f"Data Loader: The uploaded file is not a valid Excel workbook "
            f"(.xlsx / .xls).  Detected issue: {type(exc).__name__}.  "
            f"Please upload a valid .xlsx or .xls file."
        )
        return result

    # ------------------------------------------------------------------
    # Step 3: Verify Sheet 1 (data) exists; sheets 2-4 are optional.
    # ------------------------------------------------------------------
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)

    if num_sheets < 1:
        result.errors.append(
            "Data Loader: The workbook contains no sheets.  Please upload a "
            "valid MonitorQ export or a single-sheet displacement file."
        )
        wb.close()
        return result

    # Warn about missing optional sheets (V-IV Charts, Acc Charts, Summary).
    for idx, label in _SHEET_LABELS.items():
        if idx == 0:
            continue  # data sheet is mandatory — checked above
        if idx >= num_sheets:
            result.warnings.append(
                f"Data Loader: Optional sheet {label} is absent from this "
                f"workbook.  Only the displacement data will be loaded."
            )

    # ------------------------------------------------------------------
    # Step 4: Parse Sheet 1 (EnhancedDeformation).
    # ------------------------------------------------------------------
    ws_raw = wb.worksheets[0]
    raw_rows = list(ws_raw.iter_rows(values_only=True))

    if not raw_rows:
        result.errors.append(
            f"Data Loader: Sheet '{sheet_names[0]}' is empty.  "
            f"Expected a header row followed by data rows."
        )
        wb.close()
        return result

    header_row = raw_rows[0]

    # Locate the time column.  Accept any of _TIME_COLUMN_ALIASES
    # (case-insensitive, whitespace/punctuation collapsed).  This lets the
    # loader handle MonitorQ exports ('Time'), generic exports ('Timestamp',
    # 'DateTime', 'Date/Time'), and several other common variants without
    # forcing the analyst to rename the column.
    time_col_idx: int | None = None
    for col_idx, cell in enumerate(header_row):
        if _normalise_header(cell) in _TIME_COLUMN_ALIASES:
            time_col_idx = col_idx
            break

    if time_col_idx is None:
        result.errors.append(
            f"Data Loader: Sheet '{sheet_names[0]}' is missing the required "
            f"time column.  Found columns: "
            f"{[c for c in header_row if c is not None]}.  "
            f"Accepted names (case-insensitive): "
            f"{list(_TIME_COLUMN_ALIASES)}.  "
            f"Please verify the file structure."
        )
        wb.close()
        return result

    # Locate the displacement column (contains the keyword, case-insensitive).
    disp_col_idx: int | None = None
    disp_col_name: str = ""
    for col_idx, cell in enumerate(header_row):
        if (
            isinstance(cell, str)
            and _DISPLACEMENT_COLUMN_KEYWORD.lower() in cell.strip().lower()
            and col_idx != time_col_idx
        ):
            disp_col_idx = col_idx
            disp_col_name = cell.strip()
            break

    if disp_col_idx is None:
        result.errors.append(
            f"Data Loader: Sheet '{sheet_names[0]}' is missing the required "
            f"displacement column (expected a column whose name contains "
            f"'{_DISPLACEMENT_COLUMN_KEYWORD}').  Found columns: "
            f"{[c for c in header_row if c is not None]}.  "
            f"Please verify the file structure."
        )
        wb.close()
        return result

    # Build the raw DataFrame from all data rows.
    data_rows = raw_rows[1:]
    if not data_rows:
        result.errors.append(
            f"Data Loader: Sheet '{sheet_names[0]}' contains a header row "
            f"but no data rows."
        )
        wb.close()
        return result

    # Collect all column names from the header.
    col_names: list[str] = []
    for cell in header_row:
        col_names.append(str(cell) if cell is not None else "")

    raw_df = pd.DataFrame(data_rows, columns=col_names)

    # Convert the time column to pd.Timestamp.  The input may be:
    #   • an Excel serial float (legacy MonitorQ exports)
    #   • a Python datetime object (modern openpyxl with date-formatted cells)
    #   • an ISO-8601 / locale-formatted string (CSV-style exports)
    # The first parser that produces a valid Timestamp wins; everything else
    # is returned as pd.NaT and the row is dropped downstream.
    def _safe_convert_time(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return pd.NaT
        if isinstance(val, datetime):
            return pd.Timestamp(val)
        if isinstance(val, (int, float)):
            try:
                return _excel_serial_to_datetime(float(val))
            except Exception:  # noqa: BLE001
                return pd.NaT
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return pd.NaT
            # Try pandas' general parser first.
            try:
                ts = pd.to_datetime(s, errors="coerce", dayfirst=False)
                if pd.notna(ts):
                    return ts
            except Exception:  # noqa: BLE001
                pass
            # Fallback: numeric string that's actually a serial date.
            try:
                return _excel_serial_to_datetime(float(s))
            except (ValueError, TypeError):
                return pd.NaT
        return pd.NaT

    time_col_name = col_names[time_col_idx]
    raw_df[time_col_name] = raw_df[time_col_name].apply(_safe_convert_time)

    # Drop rows where the Time value could not be parsed.
    valid_time_mask = raw_df[time_col_name].notna()
    if not valid_time_mask.any():
        result.errors.append(
            f"Data Loader: Sheet '{sheet_names[0]}': no valid timestamps "
            f"could be parsed from the '{time_col_name}' column.  "
            f"Please verify the file structure."
        )
        wb.close()
        return result

    raw_df = raw_df[valid_time_mask].copy()
    raw_df = raw_df.set_index(time_col_name)
    raw_df.index = pd.DatetimeIndex(raw_df.index)
    raw_df.index.name = "Time"

    # ------------------------------------------------------------------
    # Step 5: Extract VCP names from numeric columns.
    # ------------------------------------------------------------------
    # Every non-Time numeric column is treated as a separate VCP.
    # Column name becomes the VCP identifier so the app can look up
    # raw_df[vcp_name] directly.  If only the originally found
    # displacement column is numeric, the result is a single-VCP list
    # (same behaviour as before for MonitorQ single-column exports).

    vcp_names: list[str] = []
    for col in raw_df.columns:
        raw_df[col] = pd.to_numeric(raw_df[col], errors="coerce")
        fill_ratio = raw_df[col].notna().sum() / max(len(raw_df), 1)
        if fill_ratio >= _VCP_MIN_FILL_RATIO:
            vcp_names.append(col)

    if not vcp_names:
        result.errors.append(
            "Data Loader: No monitoring points (VCPs) could be extracted "
            "from the workbook.  Expected at least one numeric displacement "
            "column (e.g. 'Enhanced Deformation', 'VCP60', 'VCP120', …)."
        )
        wb.close()
        return result

    result.vcp_names = vcp_names
    result.raw_df = raw_df

    # ------------------------------------------------------------------
    # Step 6: Emit warnings for empty chart sheets (Sheets 2 and 3).
    # ------------------------------------------------------------------
    if num_sheets > 1:
        ws_viv = wb.worksheets[1]
        if _is_sheet_empty(ws_viv):
            result.warnings.append(
                f"Data Loader: Sheet '{sheet_names[1]}' (V-IV Charts, Sheet 2) "
                f"is empty.  Chart data will not be available."
            )

    if num_sheets > 2:
        ws_acc = wb.worksheets[2]
        if _is_sheet_empty(ws_acc):
            result.warnings.append(
                f"Data Loader: Sheet '{sheet_names[2]}' (Acc Charts, Sheet 3) "
                f"is empty.  Chart data will not be available."
            )

    # ------------------------------------------------------------------
    # Step 7: Parse Sheet 4 (Summary) into summary_df.
    # ------------------------------------------------------------------
    if num_sheets > 3:
        ws_summary = wb.worksheets[3]
        summary_rows = list(ws_summary.iter_rows(values_only=True))

        if summary_rows:
            summary_header = summary_rows[0]
            summary_col_names = [
                str(c) if c is not None else f"Col_{i}"
                for i, c in enumerate(summary_header)
            ]
            summary_data = summary_rows[1:]
            summary_df = pd.DataFrame(summary_data, columns=summary_col_names)
            # Drop fully-empty rows.
            summary_df = summary_df.dropna(how="all").reset_index(drop=True)
            result.summary_df = summary_df
        else:
            result.warnings.append(
                f"Data Loader: Sheet '{sheet_names[3]}' (Summary, Sheet 4) "
                f"is empty."
            )

    wb.close()
    return result
